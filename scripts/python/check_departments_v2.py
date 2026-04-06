import openpyxl
import sys

try:
    # Load the Employee Form
    wb = openpyxl.load_workbook('Employee Form.xls')
    ws = wb.active

    with open('dept_output.txt', 'w') as f:
        f.write(f"Sheet: {ws.title}\n")
        f.write(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}\n\n")

        # Get all rows
        f.write("First 10 rows:\n")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 10:
                f.write(f"Row {i}: {row}\n")
            else:
                break

        # Extract departments - try common column positions
        f.write("\n" + "="*80 + "\n")
        f.write("Analyzing columns for departments:\n")
        f.write("="*80 + "\n\n")

        # Check columns 0-10
        for col_idx in range(min(10, ws.max_column)):
            col_values = set()
            for row_idx in range(2, min(ws.max_row + 1, 100)):  # Skip header, check up to row 100
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                if cell.value:
                    val = str(cell.value).strip()
                    if val and val.lower() not in ['nan', 'none', '']:
                        col_values.add(val)
            
            if col_values and len(col_values) <= 50:
                f.write(f"Column {col_idx} ({col_idx+1}) - {len(col_values)} unique values:\n")
                for val in sorted(col_values):
                    f.write(f"  - {val}\n")
                f.write("\n")

        f.write("Done!\n")

except Exception as e:
    with open('dept_output.txt', 'w') as f:
        f.write(f"Error: {str(e)}\n")
        f.write(f"Type: {type(e).__name__}\n")
        import traceback
        f.write(traceback.format_exc())

print("Output written to dept_output.txt")
